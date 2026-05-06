import io
import json
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from openpyxl import Workbook
from openpyxl.styles import Alignment
from app.models.test_case import TestCase
from app.models.test_point import TestPoint
from app.schemas.test_case import TestCaseUpdate, TestCaseResponse
from app.utils.exceptions import NotFoundError


async def get_test_cases(
    db: AsyncSession,
    project_id: str,
    test_point_id: str | None = None,
    priority: str | None = None,
    case_type: str | None = None,
    status: str | None = None,
    skip: int = 0,
    limit: int = 100,
) -> tuple[list[TestCaseResponse], int]:
    query = select(TestCase).where(TestCase.project_id == project_id)
    count_query = select(func.count(TestCase.id)).where(TestCase.project_id == project_id)

    if test_point_id:
        query = query.where(TestCase.test_point_id == test_point_id)
        count_query = count_query.where(TestCase.test_point_id == test_point_id)
    if priority:
        query = query.where(TestCase.priority == priority)
        count_query = count_query.where(TestCase.priority == priority)
    if case_type:
        query = query.where(TestCase.case_type == case_type)
        count_query = count_query.where(TestCase.case_type == case_type)
    if status:
        query = query.where(TestCase.status == status)
        count_query = count_query.where(TestCase.status == status)

    total = (await db.execute(count_query)).scalar()
    result = await db.execute(query.order_by(TestCase.created_at.desc()).offset(skip).limit(limit))
    cases = result.scalars().all()

    # 预加载测试点标题
    point_ids = list({c.test_point_id for c in cases})
    point_map = {}
    if point_ids:
        points_result = await db.execute(
            select(TestPoint.id, TestPoint.title).where(TestPoint.id.in_(point_ids))
        )
        point_map = {row.id: row.title for row in points_result}

    responses = []
    for c in cases:
        resp = TestCaseResponse.model_validate(c)
        resp.test_point_title = point_map.get(c.test_point_id)
        responses.append(resp)

    return responses, total


async def get_test_case(db: AsyncSession, case_id: str) -> TestCaseResponse:
    result = await db.execute(select(TestCase).where(TestCase.id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise NotFoundError("测试用例")

    # 获取测试点标题
    point_result = await db.execute(
        select(TestPoint.title).where(TestPoint.id == case.test_point_id)
    )
    point_title = point_result.scalar_one_or_none()

    resp = TestCaseResponse.model_validate(case)
    resp.test_point_title = point_title
    return resp


async def update_test_case(
    db: AsyncSession, case_id: str, data: TestCaseUpdate
) -> TestCaseResponse:
    result = await db.execute(select(TestCase).where(TestCase.id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise NotFoundError("测试用例")

    update_data = data.model_dump(exclude_unset=True)
    if "steps" in update_data and update_data["steps"]:
        update_data["steps"] = [s.model_dump() for s in data.steps]
    for key, value in update_data.items():
        setattr(case, key, value)

    await db.commit()
    await db.refresh(case)
    return TestCaseResponse.model_validate(case)


async def delete_test_case(db: AsyncSession, case_id: str) -> None:
    result = await db.execute(select(TestCase).where(TestCase.id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise NotFoundError("测试用例")

    await db.delete(case)
    await db.commit()


async def export_test_cases_excel(db: AsyncSession, project_id: str) -> bytes:
    """导出项目所有测试用例为Excel"""
    result = await db.execute(
        select(TestCase).where(TestCase.project_id == project_id)
        .order_by(TestCase.created_at)
    )
    cases = result.scalars().all()

    # 预加载测试点标题
    point_ids = list({c.test_point_id for c in cases})
    point_map = {}
    if point_ids:
        points_result = await db.execute(
            select(TestPoint.id, TestPoint.title).where(TestPoint.id.in_(point_ids))
        )
        point_map = {row.id: row.title for row in points_result}

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头
    headers = ["序号", "关联测试点", "用例标题", "前置条件", "优先级", "用例类型", "状态",
                "步骤编号", "操作步骤", "预期结果"]
    ws.append(headers)

    # 设置表头样式
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = cell.font.copy(bold=True)

    row_num = 1
    for case in cases:
        raw_steps = case.steps or []
        # 兼容双重序列化：steps 可能是 JSON 字符串而非列表
        if isinstance(raw_steps, str):
            try:
                raw_steps = json.loads(raw_steps)
            except (json.JSONDecodeError, TypeError):
                raw_steps = []
        steps = raw_steps if isinstance(raw_steps, list) else []
        point_title = point_map.get(case.test_point_id, "")

        # 将多个步骤合并到单元格内，用换行符分隔，每个用例只占一行
        step_numbers = []
        actions = []
        expected_results = []
        for i, step in enumerate(steps):
            if isinstance(step, dict):
                step_numbers.append(str(step.get("step_number", i + 1)))
                actions.append(step.get("action", ""))
                expected_results.append(step.get("expected_result", ""))
            else:
                step_numbers.append(str(i + 1))
                actions.append(str(step))
                expected_results.append("")

        ws.append([
            row_num, point_title, case.title, case.preconditions or "",
            case.priority or "", case.case_type or "", case.status,
            "\n".join(step_numbers),
            "\n".join(actions),
            "\n".join(expected_results),
        ])
        # 设置该行自动换行
        current_row = ws.max_row
        for col in range(8, 11):  # 步骤编号、操作步骤、预期结果列
            ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        row_num += 1

    # 调整列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
